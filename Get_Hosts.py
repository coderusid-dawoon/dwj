from pyVim.connect import SmartConnect
from pyVmomi import vim
import ssl
import openpyxl

# Get all the Hosts from vCenter inventory and print its name
# Below is Python 2.7.x code, which can be easily converted to python 3.x version

s=ssl.SSLContext(ssl.PROTOCOL_TLSv1)
s.verify_mode=ssl.CERT_NONE
si= SmartConnect(host="10.192.3.2", user="Administrator@vsphere.local", pwd="$h1vKamal",sslContext=None)
content=si.content

# Method that populates objects of type vimtype
def get_all_objs(content, vimtype):
    obj = {}
    container = content.viewManager.CreateContainerView(content.rootFolder, vimtype, True)
    for managed_object_ref in container.view:
        obj.update({managed_object_ref: managed_object_ref.name})
    return obj

#Calling above method
gethosts=get_all_objs(content, [vim.HostSystem],[vim.host.HardwareInfo])

# 파일 불러오기를 시도합니다.
try:
    # 워크북 불러오기
    wb = openpyxl.load_workbook("hostname.xlsx")
    sheet = wb.create_sheet()
    sheet.append(["name", "Cpu", "Memory"])
    print("불러오기 완료")
# 파일 불러오기에 실패하면, 새로운 워크북(엑셀파일)을 만듭니다.
except:
    # 워크북 새로 만들기, 현재 활성화된 시트 선택하기
    # 헤더 행 추가하기
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["name", "Cpu", "Memory"])
    print("새로운 파일을 만들었습니다.")


#Iterating each host object and printing its name
for host in gethosts:
    sheet.append([host.name])
    print(host.name)

def convertMemory(sizeBytes):
    name = ("B", "KB", "MB", "GB", "TB", "PB")
    base = int(floor(log(sizeBytes, 1024)))
    power = pow(1024,base)
    size = round(sizeBytes/power,2)
    return "{}{}".format(floor(size),name[base])

hosts = content.viewManager.CreateContainerView(content.rootFolder,[vim.HostSystem],True)
for host in hosts.view:
    # Print the hosts cpu details
    print(host.hardware.cpuInfo)
    # convert CPU to total hz to ghz times numCpuCores
    print("CPU:", round(((host.hardware.cpuInfo.hz/1e+9)*host.hardware.cpuInfo.numCpuCores),0),"GHz")
    #covert the raw bytes to readable size via convertMemory
    print("Memory:", convertMemory(host.hardware.memorySize))
