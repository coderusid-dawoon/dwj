import requests
from bs4 import BeautifulSoup
import openpyxl

# 파일 불러오기를 시도합니다.
try:
    # 워크북 불러오기
    wb = openpyxl.load_workbook("navernews.xlsx")
    sheet = wb.create_sheet()
    sheet.append(["제목", "언론사"])
    print("불러오기 완료")
# 파일 불러오기에 실패하면, 새로운 워크북(엑셀파일)을 만듭니다.
except:
    # 워크북 새로 만들기, 현재 활성화된 시트 선택하기
    # 헤더 행 추가하기
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["제목", "언론사"])
    print("새로운 파일을 만들었습니다.")

#워크시트 현재상태
sheet_names = wb.sheetnames
print(sheet_names)
# 검색어 입력받기
keyword = input("검색어를 입력해주세요: ")
#워크시트안 중복데이터 확인
i=0
while i<len(sheet_names):
    if sheet_names[i] == keyword:
        keyword = input("중복된 데이터가 있습니다. 다시 검색어를 입력해주세요: ")
        i=0
    else:
        i=i+1

# 시트 이름 변경하기
sheet.title = keyword

# 반복1: 기사번호를 변경시키면서 데이터 수집을 반복하기
# 1 ~ 100까지 10단위로 반복(1, 11, ..., 91)
for n in range(1, 100, 10):
   raw = requests.get("https://search.naver.com/search.naver?where=news&query=" + keyword + "&start=" + str(n),
                      headers={'User-Agent':'Mozilla/5.0'})
   html = BeautifulSoup(raw.text, "html.parser")
   
    # 컨테이너: ul.list_news
    # 기사제목: a.news_tit
    # 언론사: a.info.press                
   articles = html.select("ul.list_news > li")
   
   # 반복2: 기사에 대해서 반복하면 세부 정보 수집하기                
   # 리스트를 사용한 반복문으로 모든 기사에 대해서 제목/언론사 출력
   for ar in articles:
       title = ar.select_one("a.news_tit").text
       source = ar.select_one("a.info.press").text                  
       print(title, source)
       
       # 수집한 제목, 언론사를 엑셀 파일에 써줍니다.
       sheet.append([title, source])

# 작성한 워크북(엑셀파일)을 navertv.xlsx로 저장합니다.
wb.save("navernews.xlsx")